"""
workflow:
1. initiate hardware
2. generate event list for detecting surface height of stock solutions
3. run events for surface detection, get liquid surface heights and write to excel
4. generate event list for pipetting
"""

import logging


def setup_logger():
    # better logging format in console
    class CustomFormatter(logging.Formatter):
        grey = "\x1b[38;20m"
        yellow = "\x1b[33;20m"
        red = "\x1b[31;20m"
        bold_red = "\x1b[31;1m"
        reset = "\x1b[0m"
        format = "%(asctime)s - %(name)s - %(levelname)s - %(message)s (%(filename)s:%(lineno)d)"

        FORMATS = {
            logging.DEBUG: grey + format + reset,
            logging.INFO: yellow + format + reset,
            logging.WARNING: yellow + format + reset,
            logging.ERROR: red + format + reset,
            logging.CRITICAL: bold_red + format + reset
        }

        def format(self, record):
            log_fmt = self.FORMATS.get(record.levelno)
            formatter = logging.Formatter(log_fmt)
            return formatter.format(record)

    # create logger with 'main'
    logger = logging.getLogger('main')
    logger.setLevel(logging.DEBUG)
    # create file handler which logs even debug messages
    fh = logging.FileHandler('C:\\Users\\Chemiluminescence\\Dropbox\\robochem\\pipetter_files\\main.log')
    fh.setLevel(logging.INFO)
    # create console handler with a higher log level
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    # create formatter and add it to the handlers
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    ch.setFormatter(CustomFormatter())
    # add the handlers to the logger
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


logger = setup_logger()

import copy, time, pickle, re, importlib, json, os
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
import PySimpleGUI as sg
import pandas as pd
# import arrow
import zeus
import pipetter
import planner as pln
import breadboard as brb

data_folder = os.environ['ROBOCHEM_DATA_PATH'].replace('\\', '/') + '/'
# data_folder  = 'C:/Users/Chemiluminescence/Dropbox/robochem/data/'

def initiate_hardware() -> (zeus.ZeusModule, pipetter.Gantry, pipetter.Pipetter):
    # initiate zeus
    zm = zeus.ZeusModule(id=1)
    time.sleep(3)
    logger.info("zeus is loaded as: zm")

    # initiate gantry
    gt = pipetter.Gantry(zeus=zm)
    time.sleep(3)
    logger.info("gantry is loaded as: gt")
    # gt.configure_grbl() # This only need to be done once.
    gt.home_xy()
    if gt.xy_position == (0, 0):
        logger.info("gantry is homed")

    # initiate pipetter
    pt = pipetter.Pipetter(zeus=zm, gantry=gt)
    time.sleep(2)
    logger.info("pipetter is loaded as: pt")

    return zm, gt, pt

# use GUI to specify Excel path for reactions and stock solutions
def load_excel_path_by_pysimplegui():
    sg.theme('BrightColors')  # Add a touch of color
    working_directory = os.getcwd()
    # All the stuff inside your window.
    layout = [[sg.Text('Select Excel file for reactions')],
              [sg.InputText(key="-FILE_PATH-"),
               sg.FileBrowse(initial_folder=working_directory,
                             file_types=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))],
              [sg.Submit(), sg.Cancel()]]

    # Create the Window
    window = sg.Window('Select Excel file for reactions',
                       layout,
                       size=(600, 200),
                       font=('Helvetica', 14), )

    # Event Loop to process "events" and get the "values" of the inputs
    event, values = window.read()
    # print(event, values[0])
    window.close()
    logger.info(f"Excel file for reactions is selected: {values['-FILE_PATH-']}")
    return values['-FILE_PATH-']

def load_stock_solutions_from_excel(path: str) -> list:
    stock_solution_list = []
    wb_excel = load_workbook(path, data_only=True)
    ws = wb_excel[[x for x in wb_excel.sheetnames if 'Stock_solutions' in x][0]]

    for row in tuple(ws.rows)[1:]:  # exclude the header
        if row[0].value is not None:
            substance_name = row[0].value,
            index = row[1].value,
            container = row[2].value,
            solvent = row[3].value,
            density = row[4].value,
            volume = row[5].value,
            liquid_surface_height = row[6].value,
            pipetting_mode = row[7].value,
            stock_solution_list.append(
                [substance_name, index, container, solvent, density, volume, liquid_surface_height, pipetting_mode])

    # remove the tuple in the list
    for i in range(len(stock_solution_list)):
        for j in range(len(stock_solution_list[i])):
            stock_solution_list[i][j] = stock_solution_list[i][j][0]

    logger.info(f"stock solutions are loaded from Excel file: {stock_solution_list}")

    # stock_solution_list example: ['p-BrPhOTf', 'Substance_A', ' plate_4_container0', 'ethanol', 0.789, 4.5, 1979, None]

    return stock_solution_list

def generate_event_list_for_surface_detection(path_for_stock_solution: str):
    event_list_surface_detection = []

    # load object from pickle.
    # This pickle object is a template for generating events for surface detection
    pickle_path = '../calibration_for_pipetting/transfer_object_for_liquid_surface_detection.pickle'
    with open(pickle_path, 'rb') as f:
        event_for_surface_detection = pickle.load(f)

    stock_solution_list = load_stock_solutions_from_excel(path_for_stock_solution)

    # generate one event for each stock solution
    for solution in stock_solution_list:
        event_temp = copy.deepcopy(event_for_surface_detection)  # deepcopy to avoid changing the original object
        plate_id = re.findall(r'\d+', solution[2])[0]  # get the first number in the string
        container_id = re.findall(r'\d+', solution[2])[1]  # get the second number in the string
        event_temp.substance_name = solution[0]
        event_temp.source_container = copy.deepcopy(brb.plate_list[int(plate_id)].containers[int(container_id)])
        event_temp.destination_container = copy.deepcopy(event_temp.source_container)
        event_temp.asp_lld = 1  # liquidClassTableIndex =13, so pLLD will be used.
        event_temp.disp_lld = 0
        event_temp.asp_liquidClassTableIndex = 13  # use pLLD for surface detection
        event_temp.disp_liquidClassTableIndex = 13
        event_temp.asp_liquidSurface = 1200
        event_temp.asp_lldSearchPosition = zm.ZeusTraversePosition
        event_temp.disp_liquidSurface = 1800
        event_temp.aspirationVolume = 100
        event_temp.tip_type = '300ul'
        event_list_surface_detection.append(event_temp)

    return event_list_surface_detection

## This update will rely on the exact naming and order of the file header, so keep them the same as the template.
def update_excel_with_liquid_heights_and_volume(path_for_reaction: str,
                                                liquid_info_in_stock: dict):
    liquid_surface_heights_in_stock = {k[:-7]: v for k, v in liquid_info_in_stock.items() if 'height' in k}
    liquid_volume_in_stock = {k[:-7]: v for k, v in liquid_info_in_stock.items() if 'volume' in k}

    wb = load_workbook(filename=path_for_reaction)
    sheet = wb[[x for x in wb.sheetnames if 'Stock_solutions' in x][0]]

    for i in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        substance_name = i[0].value
        if substance_name in liquid_surface_heights_in_stock.keys():
            sheet[f'G{i[0].row}'] = liquid_surface_heights_in_stock[substance_name]
            sheet[f'I{i[0].row}'] = liquid_volume_in_stock[substance_name]

            logger.info(f'Stock solution {substance_name} has been updated with liquid surface height: '
                        f'{liquid_surface_heights_in_stock[substance_name]}')

    wb.save(path_for_reaction)

def add_stock_solutions_to_brb_containers(reaction_excel_path: str):
    # this loading should be done after the liquid surface heights are updated in the Excel file
    stock_solution_list = load_stock_solutions_from_excel(reaction_excel_path)
    # stock_solution_list example: ['p-BrPhOTf', 'Substance_A', ' plate_4_container0', 'ethanol', 0.789, 4.5, 1979, None]

    for solution in stock_solution_list:
        container: str = solution[2]
        plate_id = int(re.findall(r'\d+', container)[0])
        container_id = int(re.findall(r'\d+', container)[1])

        substance_name = solution[0]
        substance_index = solution[1]
        substance_solvent = solution[3]
        substance_density = solution[4]
        liquid_surface_height = solution[6]

        brb.plate_list[plate_id].add_substance_to_container(substance_name=substance_name,
                                                            container_id=container_id,
                                                            solvent=substance_solvent,
                                                            substance_density=substance_density,
                                                            liquid_surface_height=liquid_surface_height)

        brb.source_substance_containers.append(brb.plate_list[plate_id].containers[container_id])
    logger.info('All stock solutions have been added to brb.plate_list')
    return brb.source_substance_containers

if __name__ == '__main__':

    ## initiate hardware
    zm, gt, pt = initiate_hardware()

    ## load excel path
    path_for_reactions = load_excel_path_by_pysimplegui()

    ## generate event list for surface detection
    event_list_surface_detection = generate_event_list_for_surface_detection(path_for_reactions)

    ## run detection events and get liquid surface heights
    # liquid_info_in_stock example:
    # {'p-BrPhOTf_height': 1986, 'p-BrPhOTf_volume': 9.7, 'm-BrPhOTf_height': 2091, 'm-BrPhOTf_volume': 4.3}
    liquid_info_in_stock = pln.run_events_chem(zm=zm, pt=pt, logger=logger,
                                               start_event_id=0,
                                               event_list=event_list_surface_detection,
                                               prewet_tip=False)

    logger.info(f"liquid_surface_heights: {liquid_info_in_stock}")

    ## update excel with liquid surface heights
    update_excel_with_liquid_heights_and_volume(path_for_reaction = path_for_reactions, liquid_info_in_stock=liquid_info_in_stock)

    ## add stock solutions to brb containers
    add_stock_solutions_to_brb_containers(reaction_excel_path=path_for_reactions)

    # generate event list for multicomponent reactions
    #TODO:
    # 1. change the sheet name to the one you want to use
    # 2. change the usecols to the one you want to use
    event_dataframe_chem, event_list_chem = \
        pln.generate_event_object(logger=logger,
                                  excel_to_generate_dataframe=path_for_reactions,
                                  sheet_name='Reactions_0608', usecols='B:F',
                                  is_pipeting_to_balance=False, is_for_bio=False)

    # save the event list in pickle file and later load from this file
    pickle_folder = data_folder + 'multicomp-reactions\\pipetter_io\\daily_pickle_output\\'
    pickle_file = pickle_folder + f'event_list_before_run_{datetime.now().strftime("%m_%d_%H_%M")}.pickle'

    with open(pickle_file, 'wb') as f:
        pickle.dump(event_list_chem, f)

    # update planner.py when necessary
    # importlib.reload(pln)

    ## safety check
    for event in event_list_chem:
        if event.tip_type == '50ul':
            event.asp_liquidClassTableIndex = 24
            event.disp_liquidClassTableIndex = 24
        if event.tip_type == '300ul':
            event.asp_liquidClassTableIndex = 22
            event.disp_liquidClassTableIndex = 22
        if event.tip_type == '1000ul':
            event.asp_liquidClassTableIndex = 23
            event.disp_liquidClassTableIndex = 23

    # do multicomponent reactions
    pln.run_events_chem(zm=zm, pt=pt, logger=logger,
                        # event_list_path=pickle_file,
                        event_list= event_list_chem,
                        start_event_id=0,
                        prewet_tip=True)


for event in event_list_chem[174:]:
    if event.aspirationVolume >= 300:
        event.asp_liquidClassTableIndex = 23
        event.disp_liquidClassTableIndex = 23
        event.tip_type = '1000ul'

## renewal of the liquid surface after refill
for event in event_list_chem[378:]:
    event.asp_liquidSurface = 1700
    event.asp_lldSearchPosition = 1700

## the following code is cursed. do not use it....

# event_list_path = 'C:\\Users\\Chemiluminescence\\OneDrive\\roborea\\zeus-pipetter\\' \
#                   'multicomponent_reaction\\0323\\event_list_chem_0-551_events.pickle'
# with open(event_list_path, 'rb') as f:
#     event_list = pickle.load(f)
#
# for event in event_list:
#     print(event.is_event_conducted)
#     print(event.event_label)
# #
# # step1: update the event list. compare the first event to check the starting liquid surface height
# event_list_chem_later_later = copy.deepcopy(event_list_chem)
# for event in event_list_chem_later_later:
#     if event.substance_name == 'ic001':
#         print(event.event_label)
#         event.asp_liquidSurface = 1700
#         event.asp_lldSearchPosition = 1700
#
#
# ## do multicomponent reactions
# pln.run_events_chem(zm=zm, pt=pt, logger=logger,
#                     # event_list_path='multicomponent_reaction\\event_list_chem.pickle',
#                     event_list=event_list_chem_later_later,
#                     start_event_id= 0 )
#

# ## 0322 for updating DMF volume.
#
# # step1: update the event list. compare the first event to check the starting liquid surface height
# event_list_chem_later = copy.deepcopy(event_list_chem)
# for event in event_list_chem_later[1094:]:
#     if event.substance_name == 'DMF':
#         print(event.event_label)
#         event.asp_liquidSurface = event.asp_liquidSurface - 386
#         event.asp_lldSearchPosition = event.asp_liquidSurface - 50
#
# ## step2: run the event list
# pln.run_events_chem(zm=zm, pt=pt, logger=logger,
#                     # event_list_path='multicomponent_reaction\\event_list_chem.pickle',
#                     event_list=event_list_chem_later,
#                     start_event_id= 1094)
#
#
# with open('multicomponent_reaction\\event_list_chem_later.pickle', 'wb') as f:
#     pickle.dump(event_list_chem_later, f)

# with open('multicomponent_reaction\\event_list_chem.pickle', 'wb') as f:
#     pickle.dump(event_list_chem, f)
#
#
# with open('multicomponent_reaction\\event_list_chem.pickle', 'rb') as f:
#     new  = pickle.load(f)
#
# with open(event_list_path, 'rb') as f:
#     event_list = pickle.load(f)