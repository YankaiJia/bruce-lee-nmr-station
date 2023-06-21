"""
workflow:
1. initiate hardware
2. generate event list for detecting surface height of stock solutions
3. run events for surface detection, get liquid surface heights and write to excel
4. generate event list for pipetting
"""

import logging


def setup_logger():
    # better logging format in console
    class CustomFormatter(logging.Formatter):
        grey = "\x1b[38;20m"
        yellow = "\x1b[33;20m"
        red = "\x1b[31;20m"
        bold_red = "\x1b[31;1m"
        reset = "\x1b[0m"
        format = "%(asctime)s - %(name)s - %(levelname)s - %(message)s (%(filename)s:%(lineno)d)"

        FORMATS = {
            logging.DEBUG: grey + format + reset,
            logging.INFO: yellow + format + reset,
            logging.WARNING: yellow + format + reset,
            logging.ERROR: red + format + reset,
            logging.CRITICAL: bold_red + format + reset
        }

        def format(self, record):
            log_fmt = self.FORMATS.get(record.levelno)
            formatter = logging.Formatter(log_fmt)
            return formatter.format(record)

    # create logger with 'main'
    logger = logging.getLogger('main')
    logger.setLevel(logging.DEBUG)
    # create file handler which logs even debug messages
    fh = logging.FileHandler('C:\\Users\\Chemiluminescence\\Dropbox\\robochem\\pipetter_files\\main.log')
    fh.setLevel(logging.INFO)
    # create console handler with a higher log level
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    # create formatter and add it to the handlers
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    ch.setFormatter(CustomFormatter())
    # add the handlers to the logger
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


logger = setup_logger()

import numpy as np
import copy, time, pickle, re, importlib, json, os
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
import PySimpleGUI as sg
import pandas as pd
# import arrow
import zeus
import pipetter
import planner as pln
import breadboard as brb

data_folder = os.environ['ROBOCHEM_DATA_PATH'].replace('\\', '/') + '/'
# data_folder  = 'C:/Users/Chemiluminescence/Dropbox/robochem/data/'

def initiate_hardware() -> (zeus.ZeusModule, pipetter.Gantry, pipetter.Pipetter):
    # initiate zeus
    zm = zeus.ZeusModule(id=1)
    time.sleep(3)
    logger.info("zeus is loaded as: zm")

    # initiate gantry
    gt = pipetter.Gantry(zeus=zm)
    time.sleep(3)
    logger.info("gantry is loaded as: gt")
    # gt.configure_grbl() # This only need to be done once.
    gt.home_xy()
    if gt.xy_position == (0, 0):
        logger.info("gantry is homed")

    # initiate pipetter
    pt = pipetter.Pipetter(zeus=zm, gantry=gt)
    time.sleep(2)
    logger.info("pipetter is loaded as: pt")

    return zm, gt, pt

# use GUI to specify Excel path for reactions and stock solutions
def load_excel_path_by_pysimplegui():
    sg.theme('BrightColors')  # Add a touch of color
    working_directory = os.getcwd()
    # All the stuff inside your window.
    layout = [[sg.Text('Select Excel file for reactions')],
              [sg.InputText(key="-FILE_PATH-"),
               sg.FileBrowse(initial_folder=working_directory,
                             file_types=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))],
              [sg.Submit(), sg.Cancel()]]

    # Create the Window
    window = sg.Window('Select Excel file for reactions',
                       layout,
                       size=(600, 200),
                       font=('Helvetica', 14), )

    # Event Loop to process "events" and get the "values" of the inputs
    event, values = window.read()
    # print(event, values[0])
    window.close()
    logger.info(f"Excel file for reactions is selected: {values['-FILE_PATH-']}")
    return values['-FILE_PATH-']

def load_stock_solutions_from_excel(path: str) -> list:
    stock_solution_list = []
    wb_excel = load_workbook(path, data_only=True)
    ws = wb_excel[[x for x in wb_excel.sheetnames if 'stock_solutions' in x][0]]

    for row in tuple(ws.rows)[1:]:  # exclude the header
        if row[0].value is not None:
            substance_name = row[0].value,
            index = row[1].value,
            plate_id = row[2].value,
            container_id = row[3].value,
            solvent = row[4].value,
            density = row[5].value,
            volume = row[6].value,
            liquid_surface_height = row[7].value,
            pipetting_mode = row[8].value,
            stock_solution_list.append(
                {'substance_name': substance_name[0], 'index': index[0],
                 'plate_id': plate_id[0], 'container_id': container_id[0],
                 'solvent': solvent[0], 'density':density[0], 'volume': volume[0],
                 'liquid_surface_height': liquid_surface_height[0], 'pipetting_mode': pipetting_mode[0]})

    logger.info(f"stock solutions are loaded from Excel file: {stock_solution_list}")
    print(stock_solution_list)

    # stock_solution_list example:
    # {'substance_name': 'DMF', 'index': 'Substance_A', 'plate_id': 5, 'container_id': 0,
    # 'solvent': 'DMF', 'density': 0.944, 'volume': None, 'liquid_surface_height': 1804, 'pipetting_mode': 'empty'}

    return stock_solution_list

def update_stock_solution_list_to_excel(path_for_reactions: str, stock_solution_list: list):
    wb_excel = load_workbook(path_for_reactions)
    ws = wb_excel[[x for x in wb_excel.sheetnames if 'stock_solutions' in x][0]]
    for row in tuple(ws.rows)[1:]:  # exclude the header
        if row[0].value is not None:
            for solution in stock_solution_list:
                if row[0].value == solution['substance_name']:
                    row[6].value = solution['volume']
                    row[7].value = solution['liquid_surface_height']
    wb_excel.save(path_for_reactions)

def add_stock_solutions_to_containers(stock_solution_list: list) -> list:
    containers_for_stock = []

    for solution in stock_solution_list:
        solution_container = brb.plate_list[solution['plate_id']].containers[solution['container_id']]
        solution_container.substance = solution['substance_name']
        solution_container.substance_density = solution['density']
        solution_container.solvent = solution['solvent']

        solution_container.liquid_surface_height, solution_container.liquid_volume\
            = pt.check_volume_in_container(container=solution_container,liquidClassTableIndex=26,change_tip_after_each_check=True)

        containers_for_stock.append(solution_container)

    logger.info(f"Stock solutions are added to containers: {containers_for_stock}")

    return containers_for_stock

    ## safety check

def check_if_event_list_legit(event_list: list):
        for event in event_list:
            assert event.asp_liquidClassTableIndex is not None, f"asp_liquidClassTableIndex is not correct: {event.asp_liquidClassTableIndex}"
            assert event.aspirationVolume >= 0, f"aspirationVolume is not correct: {event.aspirationVolume}"
            assert event.tip_type in ['50ul', '300ul', '1000ul'], f"tip type is not correct: {event.tip_type}"
            assert event.disp_liquidClassTableIndex is not None, f"disp_liquidClassTableIndex is not correct: {event.disp_liquidClassTableIndex}"
            assert event.dispensingVolume >= 0, f"dispenseVolume is not correct: {event.dispenseVolume}"

        print("event_list is legit!")


def sort_events_according_to_aspiration_volume(event_list_chem):

        output_list = []

        # # group event by plate id
        # get mask for change of plate id
        split_index = []
        for index in range(1, len(event_list_chem)):
            if event_list_chem[index].destination_container.id['plate_id'] != \
                event_list_chem[index-1].destination_container.id['plate_id']:
                split_index.append(index)
        # group event by mask
        events_grouped_by_plate_id = np.split(event_list_chem, split_index)

        # group and sort events in each plate according to substance
        for index, plate_event in enumerate(events_grouped_by_plate_id):
            # get mask for change of substance
            split_index = []
            for index in range(1, len(plate_event)):
                if plate_event[index].substance_name != \
                    plate_event[index-1].substance_name:
                    split_index.append(index)
            # group event by mask
            split_list = np.split(plate_event, split_index)

            for event_of_one_substance in split_list:
                # sort events in each group according to aspiration volume
                sorted_list = sorted(event_of_one_substance, key=lambda x: x.aspirationVolume, reverse=True)
                output_list.extend(sorted_list)

        return output_list


if __name__ == '__main__':

    ## initiate hardware
    zm, gt, pt = initiate_hardware()

    path_for_reactions = load_excel_path_by_pysimplegui()

    stock_solution_list = load_stock_solutions_from_excel(path=path_for_reactions)

    # check the volume in stock containers and add stock solutions to containers
    containers_for_stock = add_stock_solutions_to_containers(stock_solution_list)

    # update the stock solution list
    for solution in stock_solution_list:
        for container in containers_for_stock:
            if solution['substance_name'] == container.substance:
                solution['volume'] = container.liquid_volume
                solution['liquid_surface_height'] = container.liquid_surface_height

    # update the stock solution list to Excel file
    update_stock_solution_list_to_excel(path_for_reactions=path_for_reactions, stock_solution_list=stock_solution_list)


    # generate event list for multicomponent reactions
    #TODO:
    # 1. change the sheet name to the one you want to use
    # 2. change the usecols to the one you want to use

    event_dataframe_chem, event_list_chem = \
        pln.generate_event_object(logger=logger,
                                  excel_to_generate_dataframe=path_for_reactions,
                                  is_pipeting_to_balance=False, is_for_bio=False, containers_for_stock=containers_for_stock, )


    check_if_event_list_legit(event_list_chem)

    for event in event_list_chem:
        # print(event.asp_lld)
        event.asp_lldSearchPosition -= 100

    event_list_chem_sorted = sort_events_according_to_aspiration_volume(event_list_chem)

    # save the event list in pickle file and later load from this file
    pickle_folder = data_folder + 'multicomp-reactions\\pipetter_io\\daily_pickle_output\\'
    pickle_file = pickle_folder + f'event_list_before_run_{datetime.now().strftime("%m_%d_%H_%M")}.pickle'

    with open(pickle_file, 'wb') as f:
        pickle.dump(event_list_chem_sorted, f)

    # input the sequence of plate id
    plates = input("Please input the ids of 6 plates in order: ")
    plate_code_list = [int(i) for i in plates.split()]

    if input(f'Plate id list is {plate_code_list}, continue? Y/N') not in ['Y', 'y']:
        raise Exception("The plate id list is not confirmed!")

    # use input to Y/N to confirm the event list
    if input("Run pipetting? Y/N") not in  ['Y', 'y']:
        raise Exception("The event list is not confirmed!")


    starting_id = 0
    event_for_run = event_list_chem_sorted[starting_id:]

    # do multicomponent reactions
    pln.run_events_chem(zm=zm, pt=pt, logger=logger,
                        event_list= event_for_run,
                        prewet_tip=True,
                        excel_path=path_for_reactions,
                        plate_code_list=plate_code_list)

for i in range(1360, 1370):
    print(event_list_chem_sorted[1326].substance_name)