import os
import pandas as pd
import openpyxl as xl

data_folder = os.environ['ROBOCHEM_DATA_PATH'].replace('\\', '/') + '/'

run_name = 'multicomp-reactions/2023-06-19-run01/'
destination_run = 'multicomp-reactions/2023-06-30-run01/'

#load outliers
df_known_outliers = pd.read_csv(data_folder + run_name + 'results/outliers/known_outliers.csv')
df_manual_outliers = pd.read_csv(data_folder + run_name + 'results/outliers/manual_outliers.csv')

df_all_outliers = pd.concat([df_known_outliers, df_manual_outliers, df_manual_outliers], ignore_index=True,
                            sort=False)

# randomize the order
df_all_outliers = df_all_outliers.sample(frac=1).reset_index(drop=True)

# save to excel  file certain columns
destination_excel_filename = data_folder + destination_run + f'{destination_run.split("/")[-2]}.xlsx'
columns_for_outV_excel = ['reactions', 'DMF', 'ald001', 'ptsa', 'ptsa_dil_x_5', 'am001', 'ic001']
df_all_outliers[columns_for_outV_excel].to_excel(destination_excel_filename,
                                                    index=False, sheet_name='reactions')

# copy second sheet
run_with_source_excel = 'multicomp-reactions/2023-06-20-run01/'
path1 = data_folder + run_with_source_excel + f'{run_with_source_excel.split("/")[-2]}.xlsx'
path2 = destination_excel_filename

wb1 = xl.load_workbook(filename=path1)
ws1 = wb1.worksheets[1]

wb2 = xl.load_workbook(filename=path2)
ws2 = wb2.create_sheet(ws1.title)

for row in ws1:
    for cell in row:
        ws2[cell.coordinate].value = cell.value

wb2.save(path2)